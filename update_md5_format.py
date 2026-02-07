#!/usr/bin/env python3
"""
Script to update MD5 format in existing INFO Excel files
From: {hash}|{version}
To: MD5: {hash} | Script: v{version}
"""
import hashlib
from openpyxl import load_workbook
import sys

def calculate_md5_from_row(ws, row_num, exclude_cols):
    """Calculate MD5 from a row, excluding specified columns"""
    values = []
    
    # Iterate through all columns in the row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        if cell.value is not None:
            # Find column header
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value and str(header_cell.value).strip() not in exclude_cols:
                values.append(str(cell.value).strip())
    
    # Concatenate all values and calculate MD5
    concatenated = "".join(values)
    md5_hash = hashlib.md5(concatenated.encode()).hexdigest()
    return md5_hash

def update_md5_format_in_file(file_path):
    """Update MD5 format in INFO file"""
    print(f"\nUpdating: {file_path}")
    
    try:
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find "MD5 & Script Version" column
        md5_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value and str(cell.value).strip() == "MD5 & Script Version":
                md5_col_idx = col_idx
                break
        
        if not md5_col_idx:
            print("✗ 'MD5 & Script Version' column not found!")
            return False
        
        print(f"✓ Found 'MD5 & Script Version' column at position {md5_col_idx}")
        
        # Exclude columns from MD5 calculation
        exclude_cols = ["MD5 & Script Version", "NOTE"]
        
        # Update all data rows with new MD5 format
        script_version = "3.0.0"
        row_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            # Check if row has data
            has_data = False
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value is not None:
                    has_data = True
                    break
            
            if not has_data:
                break
            
            # Calculate MD5
            md5_hash = calculate_md5_from_row(ws, row_num, exclude_cols)
            md5_value = f"MD5: {md5_hash} | Script: v{script_version}"
            
            # Set MD5 value
            ws.cell(row=row_num, column=md5_col_idx).value = md5_value
            row_count += 1
        
        print(f"✓ Updated {row_count} data rows with new format")
        
        # Save workbook
        wb.save(file_path)
        print(f"✓ Saved: {file_path}")
        
        return True
        
    except Exception as e:
        print(f"✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    files = [
        "/home/pnson/parity_generator/[INFO]_BOS_AXICRYPT.safety.xlsx",
        "/home/pnson/parity_generator/simple_test/[INFO]_SIMPLE_TOP.safety.xlsx"
    ]
    
    results = []
    for file_path in files:
        success = update_md5_format_in_file(file_path)
        results.append((file_path, success))
    
    print("\n" + "="*60)
    print("SUMMARY:")
    for file_path, success in results:
        status = "✓ SUCCESS" if success else "✗ FAILED"
        print(f"{status}: {file_path}")
