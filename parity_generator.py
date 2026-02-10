import argparse
import hashlib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from Parity_generator.common_utilities import bcolors, find_matching_port, remove_after_pattern
from Parity_generator.extract_data_classes import ExtractPort
from Parity_generator.locate_ip_classes import LocateModule, LocateInstance
from Parity_generator.generate_bus_parity import GenerateBus
from Parity_generator.remove_parity import RemoveParity
from Parity_generator.extract_info_classes import ExtractINFO, ExtractINFO_Parity_Bus
from collections import defaultdict

from Parity_generator.module_parser_utilities import CommentProcess, module_partition, module_declaration_partition
from Parity_generator.module_parser_utilities import *
from Parity_generator.declare_port import declare_parity_port_2001, declare_parity_port_1995

import warnings
import os

warnings.simplefilter(action='ignore', category=FutureWarning)

# Set AXICRYPT_HOME if not already set
if 'AXICRYPT_HOME' not in os.environ:
    os.environ['AXICRYPT_HOME'] = os.path.join(os.getcwd(), 'axicrypt')

def calculate_md5_from_info_dict(info_dict, exclude_cols=None):
    """
    Calculate MD5 hash from info_dict values.
    Excludes columns specified in exclude_cols list.
    MD5 depends only on cell values, not on their format.
    """
    if exclude_cols is None:
        exclude_cols = ["MD5 & Script Version", "NOTE"]
    
    # Collect all values except excluded columns
    values = []
    for key, value in info_dict.items():
        if key not in exclude_cols:
            # Convert to string to ensure consistent hashing
            values.append(str(value).strip())
    
    # Concatenate all values and calculate MD5
    concatenated = "".join(values)
    md5_hash = hashlib.md5(concatenated.encode()).hexdigest()
    return md5_hash

def update_info_file_with_md5(file_path, sheet_name, info_dict_list, selected_groups=None):
    """
    Update INFO Excel file with MD5 and Script Version in NOTE column.
    Format: MD5: {hash} | Script: v{version}
    """
    try:
        # Load Excel workbook
        wb = load_workbook(file_path)
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Find header row (row 2, since row 1 might be empty)
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col_idx)
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        # Find "NOTE" column
        note_col_idx = None
        for key, idx in headers.items():
            if key.startswith("NOTE"):
                note_col_idx = idx
                break
        if not note_col_idx:
            note_col_idx = ws.max_column + 1
        
        # Find "MD5 & Script Version" column
        md5_col_idx = headers.get("MD5 & Script Version")
        
        if not md5_col_idx:
            # Insert new column before NOTE
            ws.insert_cols(note_col_idx)
            md5_col_idx = note_col_idx
            # Do not set title
        
        # Exclude columns from MD5 calculation
        exclude_cols = ["MD5 & Script Version", "NOTE"]
        
        # Exclude columns from MD5 calculation
        exclude_cols = ["MD5 & Script Version", "NOTE"]
        
        # Group info_dict_list by GROUP and calculate MD5 per group
        from collections import defaultdict
        group_md5 = {}
        for info_dict in info_dict_list:
            group = info_dict.get("GROUP", "").strip()
            if group not in group_md5:
                # Calculate MD5 for the group (same for all rows in group)
                group_md5[group] = calculate_md5_from_info_dict(info_dict)
        
        # Update all data rows with MD5 values
        script_version = "3.0.0"
        
        for row_num in range(3, ws.max_row + 1):
            # Check if row has data
            has_data = False
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if cell.value is not None:
                    has_data = True
                    break
            
            if not has_data:
                break
            
            # Get GROUP from this row
            group_col_idx = headers.get("GROUP")
            if group_col_idx:
                group_cell = ws.cell(row=row_num, column=group_col_idx)
                group = str(group_cell.value).strip() if group_cell.value else ""
                
                if group in group_md5:
                    md5_hash = group_md5[group]
                    md5_value = f"MD5: {md5_hash} | Script: v{script_version}"
                    ws.cell(row=row_num, column=md5_col_idx).value = md5_value
        
        # Save workbook
        wb.save(file_path)
        print(f"✓ INFO file updated with MD5 and Script Version in MD5 column: {file_path}")
        
    except Exception as e:
        print(f"Warning: Could not update INFO file with MD5: {e}")
        print("MD5 comments will still be added to Verilog modules")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Parity generator v1 by Tho Mai (2024.06)')
    parser.add_argument('-inst', type=str, default='BOS_BUS_PARITY_AXI_M', help='Parity instance name')
    parser.add_argument('-type', type=str, default='SAFETY.PARITY', help='Parity scheme type')
    parser.add_argument("-info", type=str, help="INFO file path")
    parser.add_argument("-group", type=str, default='ALL', help="GROUP filter: comma-separated group names or 'ALL' for all groups")
    parser.add_argument("-gen-top", type=str, default='YES', help="Generate top wrapper module (YES/NO)")
    args = parser.parse_args()

    # file_path = "./Parity_generator/[INFO]_PARITY_TEMPLATE.xlsx"
    file_path = args.info
    parity_scheme_list = ["SAFETY.PARITY"]

    # Parse group filter
    if args.group.upper() == 'ALL':
        selected_groups = None  # None means all groups
    else:
        selected_groups = set([g.strip() for g in args.group.split(',')])
    
    # Parse gen_top flag
    gen_top = args.gen_top.upper() == 'YES'
    
    def extract_parity_module_ports(ip_name, Parity_generator):
        """Extract port list from generated parity module"""
        import re
        try:
            # Get parity module content
            parity_module_content = Parity_generator._generate_module_ip(ip_name)
            
            # Extract module declaration
            match = re.search(r'module\s+\w+\s*\((.*?)\);', parity_module_content, re.DOTALL)
            if match:
                port_declaration = match.group(1)
                port_names = []
                
                # Extract port names - handle multiple ports per line
                for line in port_declaration.split('\n'):
                    line = line.strip()
                    if line and not line.startswith('//'):
                        # Split by comma to handle multiple ports on same line
                        parts = line.split(',')
                        for part in parts:
                            # Extract port name (identifier after direction keywords)
                            # Remove direction keywords (input/output/inout)
                            part = re.sub(r'^\s*(input|output|inout)\s+', '', part).strip()
                            # Remove bit width specification
                            part = re.sub(r'^\[\d+.*?\]\s+', '', part).strip()
                            # Extract port name
                            match = re.search(r'(\w+)', part)
                            if match:
                                port_name = match.group(1)
                                if port_name and port_name not in port_names:
                                    port_names.append(port_name)
                
                return port_names
        except Exception as e:
            print(f"Error extracting parity module ports: {e}")
        
        return None
    
    def extract_parity_input_ports(ip_name, Parity_generator, top_port_declaration):
        """Extract input ports from parity module with their specifications (excluding ACLK, I_RESETN, and already declared ports)
        Important: Only include FIERR/ENERR control ports that are actually used for error detection, not data-related FIERR ports
        """
        import re
        try:
            parity_module_content = Parity_generator._generate_module_ip(ip_name)
            
            # Extract module declaration
            match = re.search(r'module\s+\w+\s*\((.*?)\);', parity_module_content, re.DOTALL)
            if match:
                port_declaration = match.group(1)
                extra_inport = []
                seen_ports = set()
                
                # Ports to skip (clock/reset already exist in top module)
                skip_ports = {'ACLK', 'I_CLK', 'I_RESETN', 'RESETN_ACLK'}
                # FIERR ports for data signals should NOT be included - only FIERR/ENERR for parity control
                skip_fierr_data = {'FIERR_WADDR_DATA', 'FIERR_WDATA_DATA', 'FIERR_RADDR_DATA', 'FIERR_RDATA_DATA'}
                
                # Parse each line to find input ports
                for line in port_declaration.split('\n'):
                    line = line.strip()
                    if line.startswith('input'):
                        # Split by comma to handle multiple inputs on same line
                        parts = line.split(',')
                        for part in parts:
                            part = part.strip()
                            if part.startswith('input') or part:
                                # Parse: input [width] port_name
                                match = re.search(r'input\s+(\[.*?\]\s+)?(\w+)', part)
                                if match:
                                    # Remove brackets from bit_width: "[31:0]" -> "31:0"
                                    raw_width = match.group(1).strip() if match.group(1) else ""
                                    bit_width = raw_width[1:-1] if raw_width.startswith('[') and raw_width.endswith(']') else raw_width
                                    port_name = match.group(2)
                                    
                                    # Skip if already in top module declaration, in skip list, or data-related FIERR
                                    if (port_name not in seen_ports and 
                                        port_name not in skip_ports and 
                                        port_name not in skip_fierr_data and
                                        port_name not in top_port_declaration):
                                        seen_ports.add(port_name)
                                        extra_inport.append([bit_width, port_name, ""])
                
                return extra_inport
        except Exception as e:
            print(f"Error extracting parity input ports: {e}")
        
        return []
    
    def extract_parity_output_ports(ip_name, Parity_generator, top_port_declaration):
        """Extract output ports from parity module with their specifications (excluding already declared ports)"""
        import re
        try:
            parity_module_content = Parity_generator._generate_module_ip(ip_name)
            
            # Extract module declaration
            match = re.search(r'module\s+\w+\s*\((.*?)\);', parity_module_content, re.DOTALL)
            if match:
                port_declaration = match.group(1)
                extra_outport = []
                seen_ports = set()
                
                # Parse each line to find output ports
                for line in port_declaration.split('\n'):
                    line = line.strip()
                    if line.startswith('output'):
                        # Split by comma to handle multiple outputs on same line
                        parts = line.split(',')
                        for part in parts:
                            part = part.strip()
                            if part.startswith('output') or part:
                                # Parse: output [width] port_name
                                match = re.search(r'output\s+(\[.*?\]\s+)?(\w+)', part)
                                if match:
                                    # Remove brackets from bit_width: "[31:0]" -> "31:0"
                                    raw_width = match.group(1).strip() if match.group(1) else ""
                                    bit_width = raw_width[1:-1] if raw_width.startswith('[') and raw_width.endswith(']') else raw_width
                                    port_name = match.group(2)
                                    
                                    # Skip if already in top module declaration
                                    if port_name not in seen_ports and port_name not in top_port_declaration:
                                        seen_ports.add(port_name)
                                        extra_outport.append([bit_width, port_name, ""])
                
                return extra_outport
        except Exception as e:
            print(f"Error extracting parity output ports: {e}")
        
        return []
    
    def clean_parity_from_module(module_content, module_name):
        """Remove parity-related ports and all parity instances from module"""
        import re
        
        # Remove all PARITY_GEN instances (handles nested parentheses)
        # Pattern: module_name followed by optional suffix, then instance name, then ports in parentheses
        # Handles both BOS_AXICRYPT_IP_PARITY_GEN and BOS_AXICRYPT_IP_PARITY_GEN_M variants
        parity_instance_pattern = r'\n\s*' + re.escape(module_name) + r'_IP_PARITY_GEN[_\w]*\s+u_\w+\s*\(([\s\S]*?)\)\s*;'
        
        # Use a helper function to properly match nested parentheses
        def remove_parity_instances(text):
            """Remove all parity module instances using proper bracket matching"""
            result = []
            i = 0
            while i < len(text):
                # Look for pattern: whitespace + module_name + _IP_PARITY_GEN + ... + u_ + name + (
                match = re.search(parity_instance_pattern, text[i:], re.DOTALL)
                if match:
                    # Add text before the match
                    result.append(text[i:i+match.start()])
                    # Move past the matched instance (skip it entirely)
                    i = i + match.end()
                else:
                    # No more matches, add the rest
                    result.append(text[i:])
                    break
            return ''.join(result)
        
        module_content = remove_parity_instances(module_content)
        
        # Remove lines containing parity signal names (ports and signals)
        lines = module_content.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # Skip lines that contain parity signal names (not just "PARITY" keyword)
            # Look for actual parity signal names like _PARITY or PARITY_
            if re.search(r'\w+_PARITY\b|\bPARITY_\w+', line):
                continue
            
            cleaned_lines.append(line)
        
        result = '\n'.join(cleaned_lines)
        # Final pass: remove excessive blank lines that might be left
        result = re.sub(r'\n\n\n+', '\n\n', result)
        return result
    
    def filter_by_group(info_dict_list, selected_groups):
        """Filter info_dict_list by GROUP column if selected_groups is specified"""
        if selected_groups is None:
            return info_dict_list
        
        filtered_list = []
        for info_dict in info_dict_list:
            group_value = info_dict.get("GROUP", "").strip()
            if group_value in selected_groups:
                filtered_list.append(info_dict)
        
        if not filtered_list:
            print(bcolors.WARNING + f"Warning: No rows found for groups: {selected_groups}" + bcolors.ENDC)
        
        return filtered_list

    for parity_scheme in parity_scheme_list:
        INFOExtractor = ExtractINFO(file_path, parity_scheme)
        info_dict_list = INFOExtractor._read_info_multi()
        info_dict_list = filter_by_group(info_dict_list, selected_groups)
        
        # Update INFO file with MD5 and Script Version
        if info_dict_list:
            update_info_file_with_md5(file_path, parity_scheme, info_dict_list, selected_groups)
        
        if parity_scheme == "SAFETY.PARITY":
            ip_filelist_dict = {}
            if info_dict_list:
                for info_dict in info_dict_list:
                    # Calculate MD5 from info_dict
                    md5_hash = calculate_md5_from_info_dict(info_dict)
                    
                    Parity_INFOExtractor = ExtractINFO_Parity_Bus(info_dict)
                    ip_filelist_dict[
                        Parity_INFOExtractor._extract_ip_name()] = Parity_INFOExtractor._extract_filelist_list_ip()
                    Parity_generator = GenerateBus(Parity_INFOExtractor)
                    # Set MD5 hash to class
                    GenerateBus.md5_hash = md5_hash
                    Parity_generator._wrapper_ip()

                # par_dir = f"./Parity_generator/module_parity/IP_PARITY.v"
                # par_file = open(par_dir, 'w')
                # par_file.write(Parity_generator._generate_module_ip())

                port_list = Parity_generator._list_port_ip()
                # Parity_generator._reset_generator()

                original_inport, original_outport, extra_inport, extra_outport = [], [], [], []
                if find_matching_port(original_inport, original_outport, extra_inport, extra_outport):
                    print(find_matching_port(original_inport, original_outport, extra_inport, extra_outport))
                    raise Exception("Defined extra ports already exist in design")
                

                for ip, lst in port_list.items():
                    # print("IP: ", ip)
                    original_inport, original_outport, extra_inport, extra_outport = lst
                    extra_outport_name = [outport[1] for outport in extra_outport]

                    # top_hier_wrapper
                    top_name = ip
                    file_list_list = ip_filelist_dict[ip]
                    file_list = recursive_find_v(file_list_list)
                    file_set = set(file_list)
                    file_set_env = path_env(file_set)
                    _, top_file_dir = recursive_read_v(file_set_env, top_name)
                    print(f"Target top module '{top_name}' is found at '{top_file_dir}'")
                    with open(top_file_dir, 'r') as file:
                        top_file_contents_ori = file.read()

                    # Always clean parity logic from file for processing
                    # (we'll add new parity instance if gen_top = YES)
                    top_file_contents_cleaned = clean_parity_from_module(top_file_contents_ori, top_name)

                    ModuleLocator = LocateModule(top_file_contents_cleaned, top_name)
                    before_top_file_contents, top_file_contents, after_top_file_contents = ModuleLocator._separate_ip()

                    # Add error ports to top module
                    top_file_contents_before_safety = top_file_contents
                    
                    # Extract module content for processing
                    _, module_whole_content = module_partition(top_file_contents_before_safety, top_name)

                    # For gen_top = YES: Create NEW file by:
                    # 1. Use cleaned module (old parity removed)
                    # 2. Add parity ports to declaration
                    # 3. Add parity instance before endmodule

                    if gen_top:
                        # Use cleaned module content (old parity logic removed)
                        import re as regex
                        
                        # Extract parity ports to add
                        parity_extra_inports = extract_parity_input_ports(ip, Parity_generator, top_file_contents)
                        parity_extra_outports = extract_parity_output_ports(ip, Parity_generator, top_file_contents)
                        
                        # Build parity port strings
                        parity_ports_str = ""
                        for parity_port in parity_extra_inports:
                            # parity_port[0] now contains just the range like "35-1:0" or "4:0" (no brackets)
                            port_width = ("[" + parity_port[0] + "] ") if parity_port[0] else ""
                            parity_ports_str += ",\n    input " + port_width + parity_port[1]
                        for parity_port in parity_extra_outports:
                            # parity_port[0] now contains just the range like "35-1:0" or "4:0" (no brackets)
                            port_width = ("[" + parity_port[0] + "] ") if parity_port[0] else ""
                            parity_ports_str += ",\n    output " + port_width + parity_port[1]
                        
                        # Find position to insert parity ports - right before );  at end of port list
                        # Pattern matches: optional whitespace, );, optional whitespace at end of line
                        port_list_match = regex.search(r'(\s*\)\s*;)', top_file_contents)
                        if port_list_match:
                            # Get text before the closing );
                            text_before_ending = top_file_contents[:port_list_match.start()]
                            # Remove any trailing whitespace and comma from the last port
                            text_before_ending = regex.sub(r',\s*$', '', text_before_ending)
                            # Add parity ports and close with ); on its own line
                            module_with_parity_ports = (text_before_ending + 
                                                       parity_ports_str + 
                                                       "\n);\n" +
                                                       top_file_contents[port_list_match.end():])
                        else:
                            # Fallback: just use original
                            module_with_parity_ports = top_file_contents
                        
                        # Now add parity instance before "endmodule"
                        parity_module_ports = extract_parity_module_ports(ip, Parity_generator)
                        if parity_module_ports:
                            instance_content = f"\n{top_name}_IP_PARITY_GEN u_{top_name.lower()}_ip_parity_gen ("
                            port_connections = []
                            for parity_port_name in parity_module_ports:
                                # Map clock/reset ports from top module
                                if parity_port_name == "I_CLK":
                                    top_port_name = "ACLK"
                                elif parity_port_name == "I_RESETN":
                                    top_port_name = "RESETN_ACLK"
                                else:
                                    top_port_name = parity_port_name
                                port_connections.append(f".{parity_port_name} ({top_port_name})")
                            
                            instance_content += "\n    " + ",\n    ".join(port_connections) + "\n);\n\n"
                            
                            # Find "endmodule" and insert instance before it
                            endmodule_match = regex.search(r'\bendmodule\b', module_with_parity_ports)
                            if endmodule_match:
                                endmodule_pos = endmodule_match.start()
                                top_file_contents_new = module_with_parity_ports[:endmodule_pos] + instance_content + module_with_parity_ports[endmodule_pos:]
                            else:
                                top_file_contents_new = module_with_parity_ports
                        else:
                            top_file_contents_new = module_with_parity_ports
                    else:
                        # gen_top = NO: Do not modify anything
                        top_file_contents_new = None

                    file_name = top_file_dir.split('/')[-1][:-2] if top_file_dir.endswith('.v') else \
                    top_file_dir.split('/')[-1][:-3]
                    extension = top_file_dir.split('.')[-1]
                    
                    # Generate top wrapper only if gen_top flag is set
                    if gen_top and top_file_contents_new:
                        # Create NEW file with parity ports + instance, keeping original logic
                        par_dir = remove_after_pattern(
                            "/".join(top_file_dir.split('/')[:-1])) + f"/SAFETY/{file_name}_NEW.{extension}"
                        par_file = open(par_dir, 'w')
                        par_file.write(before_top_file_contents + top_file_contents_new + after_top_file_contents)
                        par_file.close()
                        print(bcolors.OKGREEN + f"Finished creating top module wrapper {par_dir}" + bcolors.ENDC)
                    # Note: When gen_top = NO, original file is NOT modified. Only parity module is generated.

                    par_dir = remove_after_pattern(
                        "/".join(top_file_dir.split('/')[:-1])) + f"/SAFETY/{file_name}_PARITY_NEW.{extension}"
                    par_file = open(par_dir, 'w')
                    par_file.write(Parity_generator._generate_module_ip(ip))
                    print(bcolors.OKGREEN + f"Finished creating parity module {par_dir}" + bcolors.ENDC)
                Parity_generator._reset_generator()

                Parity_generator._reset_generator()

        else:
            raise ValueError("Invalid Parity scheme")
