#!/usr/bin/env python3
"""
Script to add "MD5 & Script Version" column to INFO Excel files
"""
import hashlib
from openpyxl import load_workbook
import sys

def calculate_md5_from_row(ws, row_num, exclude_cols):
    """Calculate MD5 from a row, excluding specified columns"""
    values = []
    
    # Iterate through all columns in the row
    for cell in ws[row_num]:
        if cell.value is not None:
            col_letter = cell.column_letter
            # Find column header
            header_cell = ws[f'{col_letter}1']
            if header_cell.value and str(header_cell.value).strip() not in exclude_cols:
                values.append(str(cell.value).strip())
    
    # Concatenate all values and calculate MD5
    concatenated = "".join(values)
    md5_hash = hashlib.md5(concatenated.encode()).hexdigest()
    return md5_hash

def add_md5_column_to_file(file_path):
    """Add MD5 & Script Version column to INFO file"""
    print(f"\nProcessing: {file_path}")
    
    try:
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find header row (row 1)
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        
        print(f"Found columns: {list(headers.keys())}")
        
        # Find "NOTE" column
        note_col_idx = headers.get("NOTE")
        if not note_col_idx:
            print("WARNING: 'NOTE' column not found!")
            note_col_idx = max(headers.values()) + 1
        
        # Find or create "MD5 & Script Version" column
        md5_col_idx = headers.get("MD5 & Script Version")
        
        if not md5_col_idx:
            # Insert new column before NOTE
            ws.insert_cols(note_col_idx)
            md5_col_idx = note_col_idx
            ws.cell(row=1, column=md5_col_idx).value = "MD5 & Script Version"
            print(f"✓ Inserted 'MD5 & Script Version' column at position {md5_col_idx}")
        else:
            print(f"✓ 'MD5 & Script Version' column already exists at position {md5_col_idx}")
        
        # Exclude columns from MD5 calculation
        exclude_cols = ["MD5 & Script Version", "NOTE"]
        
        # Update all data rows with MD5 values
        script_version = "3.0.0"
        row_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            # Check if row has data
            has_data = False
            for cell in ws[row_num]:
                if cell.value is not None:
                    has_data = True
                    break
            
            if not has_data:
                break
            
            # Calculate MD5
            md5_hash = calculate_md5_from_row(ws, row_num, exclude_cols)
            md5_value = f"MD5: {md5_hash} | Script: v{script_version}"
            
            # Set MD5 value
            ws.cell(row=row_num, column=md5_col_idx).value = md5_value
            row_count += 1
        
        print(f"✓ Updated {row_count} data rows with MD5 values")
        
        # Save workbook
        wb.save(file_path)
        print(f"✓ Saved: {file_path}")
        
        return True
        
    except Exception as e:
        print(f"✗ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    files = [
        "/home/pnson/parity_generator/[INFO]_BOS_AXICRYPT.safety.xlsx",
        "/home/pnson/parity_generator/simple_test/[INFO]_SIMPLE_TOP.safety.xlsx"
    ]
    
    results = []
    for file_path in files:
        success = add_md5_column_to_file(file_path)
        results.append((file_path, success))
    
    print("\n" + "="*60)
    print("SUMMARY:")
    for file_path, success in results:
        status = "✓ SUCCESS" if success else "✗ FAILED"
        print(f"{status}: {file_path}")
